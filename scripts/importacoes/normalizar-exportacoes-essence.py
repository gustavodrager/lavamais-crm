#!/usr/bin/env python3
"""Converte XLSX do Essence nos CSVs aceitos pelo ImportadorEssence."""
from __future__ import annotations
import argparse,csv,hashlib,json,re
from collections import defaultdict
from datetime import datetime,timezone,timedelta
from pathlib import Path
from typing import Any
from openpyxl import load_workbook

FUSO=timezone(timedelta(hours=-3))
def txt(v): return re.sub(r"\s+"," ",str(v or "").strip())
def digs(v): return re.sub(r"\D","",txt(v))
def fone(v): return {digs(x) for x in re.findall(r"\d[\d\s().+-]{8,20}",txt(v)) if len(digs(x)) in (10,11)}
def valor(v):
    if isinstance(v,(int,float)): return float(v)
    s=txt(v).replace("R$","").replace(".","").replace(",",".").strip()
    return 0.0 if s in ("","-") else float(s)
def data(d,h=None):
    s=f"{txt(d)} {txt(h)}".strip()
    for fmt in ("%d/%m/%Y %H:%M","%d/%m/%Y %H:%M:%S","%d/%m/%Y"):
        try:return datetime.strptime(s,fmt).replace(tzinfo=FUSO)
        except ValueError:pass
    raise ValueError(f"Data invalida: {s}")
def linhas(path): return list(load_workbook(path,read_only=True,data_only=True).active.iter_rows(values_only=True))
def tabela(path,cab,inicio):
    rs=linhas(path); hs=[txt(v) for v in rs[cab]]
    return [dict(zip(hs,r)) for r in rs[inicio:] if any(v is not None for v in r)]
def hash_arq(path): return hashlib.sha256(path.read_bytes()).hexdigest()
def escrever(path,campos,rs):
    with path.open("w",encoding="utf-8",newline="") as f:
        w=csv.DictWriter(f,fieldnames=campos); w.writeheader(); w.writerows(rs)

def main():
    p=argparse.ArgumentParser()
    for a in ("movimentacoes","itens","historico","produtos"): p.add_argument(f"--{a}",required=True,type=Path)
    p.add_argument("--saida",required=True,type=Path); o=p.parse_args(); o.saida.mkdir(parents=True,exist_ok=True)
    mov=tabela(o.movimentacoes,0,1); itens=tabela(o.itens,3,4); hist=tabela(o.historico,3,5); prod=tabela(o.produtos,2,3)
    ident={}
    def add(c,n,t,dt):
        if not c:return
        x=ident.setdefault(c,{"nomes":set(),"fones":set(),"ultimo":None}); x["nomes"].add(txt(n)); fs=fone(t); x["fones"].update(fs)
        if fs and (x["ultimo"] is None or dt>x["ultimo"][0]): x["ultimo"]=(dt,sorted(fs)[0])
    for r in mov:add(digs(r["Código Cliente"]),r["Cliente"],r["Telefone"],data(r["Data Lançamento"],r["Horário Lançamento"]))
    for r in hist:add(digs(r["Código"]),r["Cliente"],r["Telefone"],data(r["Data Lançamento"]))
    cods=defaultdict(set)
    for c,x in ident.items():
        for f in x["fones"]:cods[f].add(c)
    aceitos={}; pend=[]
    for c,x in sorted(ident.items()):
        motivos=[]
        if not x["fones"]:motivos.append("telefone_ausente_ou_invalido")
        if len(x["fones"])>1:motivos.append("multiplos_telefones")
        if any(len(cods[f])>1 for f in x["fones"]):motivos.append("telefone_compartilhado")
        if len(x["nomes"])>1:motivos.append("nomes_divergentes")
        if motivos:pend.append({"codigoExterno":c,"motivos":sorted(set(motivos))})
        else:aceitos[c]={"codigo_externo":c,"nome":max(x["nomes"],key=len),"whatsapp":next(iter(x["fones"]))}
    saida_mov=[]; sem=0
    def append(r,historico=False):
        nonlocal sem
        c=digs(r["Código"] if historico else r["Código Cliente"])
        if c not in aceitos:sem+=1;return
        total=valor(r["Valor Ticket"] if historico else r["Total Ticket"])
        bruto=txt(r["Ticket"]); ticket=(f"COLETA-{digs(bruto)}" if historico and "COLETA" in bruto.upper() else digs(bruto))
        saida_mov.append({"ticket":ticket,"codigo_cliente":c,"data_hora":data(r["Data Lançamento"],None if historico else r["Horário Lançamento"]).isoformat(timespec="seconds"),"pecas":1 if historico else int(r["Peças"]),"total":f"{total:.2f}","subtotal":f"{(total if historico else valor(r['SubTotal Ticket'])):.2f}","desconto":"0.00" if historico else f"{valor(r['Desconto']):.2f}","pacote":"false" if historico else ("true" if txt(r["Ticket de Pacote"]).lower()=="sim" else "false"),"valor_utilizacao_pacote":"" if historico or txt(r["Valor Médio Utilização de Pacote"])=="-" else f"{valor(r['Valor Médio Utilização de Pacote']):.2f}","atendente":"" if historico else txt(r["Vendedor"])})
    for r in hist:append(r,True)
    for r in mov:append(r)
    produtos_agrupados={}
    for r in prod:
        q=r["QUANTIDADE"]
        if not isinstance(q,(int,float)):
            m=re.search(r"\((\d+) itens\)",txt(q))
            if not m:continue
            q=int(m.group(1))
        nome=txt(r["PRODUTO"]); chave=nome.upper()
        agregado=produtos_agrupados.setdefault(chave,{"nome":nome,"quantidade":0,"total":0.0,"periodo_inicio":"2026-01-01","periodo_fim":"2026-09-03"})
        agregado["quantidade"]+=int(q); agregado["total"]+=valor(r["TOTAL"])
    produtos=[]
    for agregado in produtos_agrupados.values():
        agregado["total"]=f"{agregado['total']:.2f}"; produtos.append(agregado)
    por_ticket=defaultdict(int); ids=set()
    for r in itens:por_ticket[digs(r["Sequência"])]+=int(r["Quant. Item"]);ids.add(txt(r["Ident. Item"]))
    div=[digs(r["Ticket"]) for r in mov if digs(r["Ticket"]) in por_ticket and por_ticket[digs(r["Ticket"])]!=int(r["Peças"])]
    escrever(o.saida/"clientes.csv",["codigo_externo","nome","whatsapp"],aceitos.values())
    escrever(o.saida/"movimentacoes.csv",["ticket","codigo_cliente","data_hora","pecas","total","subtotal","desconto","pacote","valor_utilizacao_pacote","atendente"],saida_mov)
    escrever(o.saida/"produtos.csv",["nome","quantidade","total","periodo_inicio","periodo_fim"],produtos)
    rel={"modo":"normalizacao_sem_gravacao","arquivos":{k:{"nome":v.name,"sha256":hash_arq(v)} for k,v in {"movimentacoes":o.movimentacoes,"itens":o.itens,"historico":o.historico,"produtos":o.produtos}.items()},"clientes":{"observados":len(ident),"aceitos":len(aceitos),"pendentes":len(pend)},"movimentacoes":{"observadas":len(hist)+len(mov),"aceitas":len(saida_mov),"sem_cliente_conciliado":sem,"valor_aceito":round(sum(float(x["total"]) for x in saida_mov),2)},"itens":{"observados":len(itens),"identificadores_unicos":len(ids),"tickets":len(por_ticket),"divergencias_de_pecas":div,"destino":"somente_reconciliacao_conforme_ADR_016"},"produtos":{"observados":len(prod),"normalizados":len(produtos),"linhas_consolidadas":len(prod)-len(produtos)},"pendenciasDeIdentidade":pend}
    (o.saida/"normalizacao.json").write_text(json.dumps(rel,ensure_ascii=False,indent=2),encoding="utf-8")
    print(json.dumps({k:v for k,v in rel.items() if k!="pendenciasDeIdentidade"},ensure_ascii=False,indent=2))
if __name__=="__main__":main()
